"""
Walk Audit PDF Scraper
======================
Downloads walk audit PDFs from a Google Sheet and extracts street-level photos
from each PDF. A heuristic filter (`is_likely_logo`) discards non-photo images
(logos, icons, letterhead graphics) before saving.

Logo-filtering heuristics
--------------------------
An image is treated as a logo/non-photo and skipped if ANY of the following
are true:

1. **Too small** - width < MIN_IMAGE_WIDTH or height < MIN_IMAGE_HEIGHT.
   Real photos are large; logos and icons are typically small.

2. **Extreme aspect ratio** - width/height > MAX_ASPECT_RATIO or
   width/height < MIN_ASPECT_RATIO.  Banners and tall narrow graphics are
   rarely street photos.

3. **Has transparency** - any pixel with alpha < 255.  Logos are often saved
   with transparent backgrounds; street photos almost never are.

4. **Too few unique colors** - distinct colors ≤ MIN_UNIQUE_COLORS.
   Photos contain thousands of subtly different tones; flat graphics use a
   small palette.

5. **Nearly all white (overall)** - average channel brightness > 240.
   Blank or near-blank pages get caught here.

6. **All-white corners** - all four corner pixels have average brightness
   ≥ MAX_EDGE_BRIGHTNESS.  Logos placed on white backgrounds typically leave
   white corners; real street photos rarely do.

Configuration
-------------
Set FRESH_DOWNLOAD = True to re-download all PDFs from the spreadsheet.
Set FRESH_DOWNLOAD = False (default) to re-extract images from already-
downloaded PDFs without hitting the network again.
"""

import io
import os
import re
import shutil
import time

import fitz  # PyMuPDF
import requests
from openpyxl import load_workbook
from PIL import Image, ImageOps, ImageStat
from rich.console import Console
from rich.panel import Panel
from rich.progress import BarColumn, MofNCompleteColumn, Progress, TextColumn

console = Console()

# --- Configuration ---
FRESH_DOWNLOAD = True
OUTPUT_DIR = "data/download"
FIRST_N_ROWS = None

# --- Image Filtering Configuration ---
MIN_IMAGE_WIDTH = 250
MIN_IMAGE_HEIGHT = 200
MAX_ASPECT_RATIO = 3.0
MIN_ASPECT_RATIO = 0.33
MIN_UNIQUE_COLORS = 1500
MAX_EDGE_BRIGHTNESS = 250


def is_likely_logo(image_bytes):
    """Analyzes image bytes to filter out logos, icons, and flat graphics."""
    try:
        img = Image.open(io.BytesIO(image_bytes))
        width, height = img.size
        if width < MIN_IMAGE_WIDTH or height < MIN_IMAGE_HEIGHT:
            return True

        aspect_ratio = width / height
        if aspect_ratio > MAX_ASPECT_RATIO or aspect_ratio < MIN_ASPECT_RATIO:
            return True

        if img.mode in ("RGBA", "LA") or (
            img.mode == "P" and "transparency" in img.info
        ):
            alpha = img.convert("RGBA").getchannel("A")
            alpha_extrema = alpha.getextrema()  # (min_alpha, max_alpha)
            if alpha_extrema and alpha_extrema[0] < 255:
                return True

        rgb_img = img.convert("RGB")
        colors = rgb_img.getcolors(maxcolors=MIN_UNIQUE_COLORS + 1)
        if colors is not None and len(colors) <= MIN_UNIQUE_COLORS:
            return True

        stat = ImageStat.Stat(rgb_img)
        avg_brightness = sum(stat.mean) / 3
        if avg_brightness > 240:
            return True

        corners = [
            rgb_img.getpixel((0, 0)),
            rgb_img.getpixel((width - 1, 0)),
            rgb_img.getpixel((0, height - 1)),
            rgb_img.getpixel((width - 1, height - 1)),
        ]
        if all(sum(c) / 3 >= MAX_EDGE_BRIGHTNESS for c in corners):
            return True

        return False
    except Exception:
        return True


def sanitize_folder_name(name):
    if name is None:
        return "unknown"
    name = str(name).split(",")[0].strip()
    try:
        if isinstance(name, (int, float)) or (
            isinstance(name, str) and name.strip().replace(".", "", 1).isdigit()
        ):
            name = str(int(float(name)))
        else:
            name = str(name).lower()
    except (ValueError, TypeError):
        name = str(name).lower()
    name = re.sub(r"[^a-z0-9]+", "_", name)
    return name.strip("_")


def parse_city_neighborhood(raw):
    """Split 'Boston (East)' into ('boston', 'east'), or 'Springfield' into ('springfield', None)."""
    if raw is None:
        return "unknown", None
    raw = str(raw).strip()
    match = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", raw)
    if match:
        city = sanitize_folder_name(match.group(1).strip())
        neighborhood = sanitize_folder_name(match.group(2).strip())
        return city, neighborhood
    return sanitize_folder_name(raw), None


def _slug(text):
    """Lowercase slug with hyphens replacing any run of non-alphanumeric characters."""
    if text is None:
        return "unknown"
    text = str(text).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def get_audit_identifier(city_raw, year_raw, street_raw):
    """Build a filesystem-safe identifier for a walk audit record.

    Format:  CITY_YEAR_STREETS
             CITY_YEAR_NEIGHBORHOOD_STREETS   (when neighborhood present)

    Underscores separate the four structural parts; hyphens replace spaces
    within each part.  Only the first 4 words of the street name are used.

    Examples:
        "Springfield", 2013, "Marshall Street"   → "springfield_2013_marshall-street"
        "Boston (East)", 2016, "Maverick Square"  → "boston_2016_east_maverick-square"
    """
    raw = str(city_raw).strip() if city_raw else ""
    match = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", raw)
    if match:
        city = _slug(match.group(1))
        neighborhood = _slug(match.group(2))
    else:
        city = _slug(raw)
        neighborhood = None

    year = str(int(float(year_raw))) if year_raw is not None else "unknown"
    street = "-".join(_slug(street_raw).split("-")[:4])

    parts = [city, year]
    if neighborhood:
        parts.append(neighborhood)
    parts.append(street)
    return "_".join(parts)


def get_direct_download_url(url):
    if not url or not isinstance(url, str):
        return url
    if "drive.google.com" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            return f"https://drive.google.com/uc?export=download&id={match.group(1)}"
    if "docs.google.com/document" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            return (
                f"https://docs.google.com/document/d/{match.group(1)}/export?format=pdf"
            )
    return url


def crop_caption_bands(img):
    """Crop uniform dark bands from the top/bottom of an image.

    Targets the black-box artifact left when a PDF caption text layer sat on
    top of a reserved dark region in the embedded image.  Scans inward in
    small bands; stops as soon as a band is bright or varied enough to be
    real photo content.  Never crops more than 25 % from either edge so a
    legitimately dark photo isn't destroyed.
    """
    gray = img.convert("L")
    width, height = gray.size
    BAND = 8
    MAX_MEAN = 25  # brightness threshold (0-255) — below this = "black"
    MAX_STD = 15  # std-dev threshold — below this = "uniform"
    MAX_CROP = 0.25  # never crop more than this fraction of image height

    def is_dark_uniform(y0):
        region = gray.crop((0, y0, width, min(y0 + BAND, height)))
        stat = ImageStat.Stat(region)
        return stat.mean[0] < MAX_MEAN and stat.stddev[0] < MAX_STD

    top = 0
    while top + BAND <= height * MAX_CROP and is_dark_uniform(top):
        top += BAND

    bottom = height
    while bottom - BAND >= height * (1 - MAX_CROP) and is_dark_uniform(bottom - BAND):
        bottom -= BAND

    if top == 0 and bottom == height:
        return img
    return img.crop((0, top, width, bottom))


def correct_image_rotation(image_bytes, ext, page_rotation=0):
    """Fix rotation via EXIF metadata and PDF page-level rotation.

    Returns (corrected_bytes, was_rotated).
    """
    try:
        img = Image.open(io.BytesIO(image_bytes))

        exif_orientation = img.getexif().get(274, 1)  # tag 274 = Orientation
        exif_rotated = exif_orientation != 1

        img = ImageOps.exif_transpose(img)
        if page_rotation:
            img = img.rotate(page_rotation, expand=True)

        img = crop_caption_bands(img)

        was_rotated = exif_rotated or bool(page_rotation)
        buf = io.BytesIO()
        save_fmt = img.format or ("JPEG" if ext.lower() in ("jpg", "jpeg") else "PNG")
        img.save(buf, format=save_fmt)
        return buf.getvalue(), was_rotated
    except Exception:
        return image_bytes, False


def _placement_rotation(page, xref):
    """Return the rotation (0/90/180/270) encoded in the image's placement matrix.

    PyMuPDF extracts raw pixel bytes without applying the page's content-stream
    transform. When a PDF embeds an image sideways (e.g. a portrait photo placed
    as landscape), the rotation lives in the placement matrix, not in the page
    rotation field or the image's EXIF. We detect it from the matrix shape:

      Normal   Matrix(w,  0,  0, h, …)  → 0°
      90° CW   Matrix(0, -h,  w, 0, …)  → 270° CCW to correct
      180°     Matrix(-w, 0,  0,-h, …)  → 180°
      90° CCW  Matrix(0,  h, -w, 0, …)  → 90° CCW to correct
    """
    rects = page.get_image_rects(xref, transform=True)
    if not rects:
        return 0
    m = rects[0][1]
    a, b = m.a, m.b
    if abs(b) < 1e-3:  # no off-diagonal → 0° or 180°
        return 180 if a < 0 else 0
    if abs(a) < 1e-3:  # off-diagonal dominant → 90° or 270°
        return 90 if b > 0 else 270
    return 0


def extract_images_from_pdf(pdf_path, output_subfolder, progress=None, task=None):
    if not os.path.exists(pdf_path):
        return 0
    image_folder = os.path.join(output_subfolder, "images")
    if os.path.exists(image_folder):
        shutil.rmtree(image_folder)

    try:
        doc = fitz.open(pdf_path)
        image_count = 0
        rotated_count = 0
        for i in range(len(doc)):
            page = doc[i]
            page_rotation = page.rotation
            for img in page.get_images():
                xref = img[0]
                matrix_rotation = _placement_rotation(page, xref)
                total_rotation = (page_rotation + matrix_rotation) % 360
                base_image = doc.extract_image(xref)
                image_data, was_rotated = correct_image_rotation(
                    base_image["image"], base_image["ext"], total_rotation
                )
                if is_likely_logo(image_data):
                    continue

                if not os.path.exists(image_folder):
                    os.makedirs(image_folder)
                image_count += 1
                if was_rotated:
                    rotated_count += 1
                with open(
                    os.path.join(
                        image_folder, f"image_{image_count}.{base_image['ext']}"
                    ),
                    "wb",
                ) as f:
                    f.write(image_data)
        doc.close()
        return image_count, rotated_count
    except Exception as e:
        console.print(f"    [red][Images Error][/red] {pdf_path}: {e}")
        return 0, 0


def download_pdf(url, folder_path, filename, retries=2):
    if not url or url == "No Link Found" or not url.startswith("http"):
        return False
    download_url = get_direct_download_url(url)
    for attempt in range(1 + retries):
        try:
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            file_path = os.path.join(folder_path, filename)
            response = requests.get(download_url, timeout=45, stream=True)
            response.raise_for_status()
            with open(file_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            return True
        except Exception as e:
            if attempt < retries:
                wait = 2**attempt  # 1s, then 2s
                console.print(
                    f"  [yellow][Retry {attempt + 1}/{retries}][/yellow] {url} — waiting {wait}s"
                )
                time.sleep(wait)
            else:
                console.print(f"  [red][Error][/red] {url}: {e}")
    return False


def fetch_and_extract(sheet_id, gid):
    console.print(
        Panel.fit(
            "[bold cyan]Walk Audit Scraper[/bold cyan]\n"
            f"Mode: [yellow]{'Fresh download' if FRESH_DOWNLOAD else 'Re-extract from existing PDFs'}[/yellow]",
            border_style="cyan",
        )
    )

    if FRESH_DOWNLOAD:
        if os.path.exists(OUTPUT_DIR):
            shutil.rmtree(OUTPUT_DIR)
        os.makedirs(OUTPUT_DIR)

        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid={gid}"
        excel_path = os.path.join(OUTPUT_DIR, "walk_audit_data.xlsx")

        try:
            console.print("[dim]Fetching spreadsheet...[/dim]")
            res = requests.get(export_url)
            with open(excel_path, "wb") as f:
                f.write(res.content)
            wb = load_workbook(excel_path)
            ws = wb.active

            # Locate columns by header name (Row 2)
            headers = ws[2]

            def find_col_index(substring):
                for i, cell in enumerate(headers):
                    if cell.value and substring in str(cell.value).upper():
                        return i
                return None

            idx_city = find_col_index("CITY")
            idx_year = find_col_index("YEAR")
            idx_streets = find_col_index("STREETS")
            idx_view = find_col_index("VIEW")

            if None in (idx_city, idx_year, idx_streets, idx_view):
                missing = []
                if idx_city is None:
                    missing.append("CITY")
                if idx_year is None:
                    missing.append("YEAR")
                if idx_streets is None:
                    missing.append("STREETS")
                if idx_view is None:
                    missing.append("VIEW")
                console.print(
                    f"[red][Error][/red] Missing columns: {', '.join(missing)}"
                )
                return

            rows = [row for row in ws.iter_rows(min_row=3) if row[idx_city].value]
            if FIRST_N_ROWS:
                rows = rows[:FIRST_N_ROWS]

            with Progress(
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                MofNCompleteColumn(),
                console=console,
            ) as progress:
                task = progress.add_task("[cyan]Downloading audits...", total=len(rows))
                for row in rows:
                    folder_name = get_audit_identifier(
                        row[idx_city].value, row[idx_year].value, row[idx_streets].value
                    )
                    link_cell = row[idx_view]
                    link_url = (
                        link_cell.hyperlink.target
                        if link_cell.hyperlink
                        else "No Link Found"
                    )
                    progress.update(task, description=f"[cyan]{folder_name}[/cyan]")
                    folder_path = os.path.join(OUTPUT_DIR, folder_name)
                    ok = download_pdf(link_url, folder_path, "audit.pdf")
                    if ok:
                        n, r = extract_images_from_pdf(
                            os.path.join(folder_path, "audit.pdf"), folder_path
                        )
                        rotation_note = f", [yellow]{r} rotated[/yellow]" if r else ""
                        console.log(
                            f"[green]✓[/green] {folder_name} "
                            f"[dim]({n} image{'s' if n != 1 else ''}{rotation_note})[/dim]"
                        )
                    else:
                        console.log(
                            f"[yellow]–[/yellow] {folder_name} [dim](no PDF)[/dim]"
                        )
                    progress.advance(task)

        except Exception as e:
            console.print(f"[red]Error:[/red] {e}")

    else:
        all_subdirs = [
            f
            for f in os.listdir(OUTPUT_DIR)
            if os.path.isdir(os.path.join(OUTPUT_DIR, f))
        ]
        for folder_name in all_subdirs:
            images_path = os.path.join(OUTPUT_DIR, folder_name, "images")
            if os.path.exists(images_path):
                shutil.rmtree(images_path)

        folders = [
            f
            for f in all_subdirs
            if os.path.exists(os.path.join(OUTPUT_DIR, f, "audit.pdf"))
        ]
        with Progress(
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            MofNCompleteColumn(),
            console=console,
        ) as progress:
            task = progress.add_task(
                "[cyan]Re-extracting images...", total=len(folders)
            )
            for folder_name in folders:
                folder_path = os.path.join(OUTPUT_DIR, folder_name)
                progress.update(task, description=f"[cyan]{folder_name}[/cyan]")
                n, r = extract_images_from_pdf(
                    os.path.join(folder_path, "audit.pdf"), folder_path
                )
                rotation_note = f", [yellow]{r} rotated[/yellow]" if r else ""
                console.log(
                    f"[green]✓[/green] {folder_name} "
                    f"[dim]({n} image{'s' if n != 1 else ''}{rotation_note})[/dim]"
                )
                progress.advance(task)

    console.print("[bold green]Done.[/bold green]")


if __name__ == "__main__":
    fetch_and_extract("1-Vxf7AlXk_WJwwYSVy7F28qjxVXQOAmQ-NN0JImx95Y", "379989993")
