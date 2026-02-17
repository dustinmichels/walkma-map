import os
import re
import shutil

import fitz  # PyMuPDF
import requests
from openpyxl import load_workbook

# Global configuration
OUTPUT_DIR = "data/download"
PRIMARY_IMAGE_DIR = "app/public/data/images"
FIRST_N_ROWS = 5  # Set to None to process all rows, or an integer to limit processing

# Image filtering configuration
MIN_IMAGE_WIDTH = 200  # Minimum width in pixels to keep an image (filters logos)
MIN_IMAGE_HEIGHT = 200  # Minimum height in pixels to keep an image (filters icons)


def sanitize_folder_name(name):
    """
    Converts strings like 'Barnstable (Hyannis)' into 'barnstable_hyannis'.
    Cleans numeric values to prevent trailing decimals/underscores (e.g. 2018.0 -> 2018).
    """
    if name is None:
        return "unknown"

    # Try to convert to int to drop decimals (e.g., 2018.0 -> 2018)
    try:
        # If it's a number or a string that looks like a number
        if isinstance(name, (int, float)) or (
            isinstance(name, str) and name.strip().isdigit()
        ):
            name = str(int(float(name)))
        else:
            name = str(name).lower()
    except (ValueError, TypeError):
        name = str(name).lower()

    # Replace non-alphanumeric characters with underscores
    name = re.sub(r"[^a-z0-9]+", "_", name)
    # Remove leading/trailing underscores
    return name.strip("_")


def get_direct_download_url(url):
    """
    Converts Google Drive, Google Docs, or Google Sheets links into direct PDF download/export links.
    """
    if not url or not isinstance(url, str):
        return url

    # 1. Handle standard Google Drive file links
    if "drive.google.com" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            file_id = match.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"

    # 2. Handle Google Docs links -> Export as PDF
    if "docs.google.com/document" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            file_id = match.group(1)
            return f"https://docs.google.com/document/d/{file_id}/export?format=pdf"

    # 3. Handle Google Sheets links -> Export as PDF
    if "docs.google.com/spreadsheets" in url:
        match = re.search(r"/d/([^/]+)", url)
        if match:
            file_id = match.group(1)
            return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=pdf"

    return url


def extract_images_from_pdf(pdf_path, output_subfolder):
    """
    Extracts all images from a PDF file and saves them to a subfolder.
    Filters out small images (logos/icons) based on pixel dimensions.
    """
    if not os.path.exists(pdf_path):
        return

    image_folder = os.path.join(output_subfolder, "images")

    try:
        doc = fitz.open(pdf_path)
        image_count = 0
        skipped_count = 0

        for i in range(len(doc)):
            for img_index, img in enumerate(doc.get_page_images(i)):
                xref = img[0]
                base_image = doc.extract_image(xref)

                # Check dimensions to filter out logos
                width = base_image.get("width", 0)
                height = base_image.get("height", 0)

                if width < MIN_IMAGE_WIDTH or height < MIN_IMAGE_HEIGHT:
                    skipped_count += 1
                    continue

                image_bytes = base_image["image"]
                image_ext = base_image["ext"]

                if not os.path.exists(image_folder):
                    os.makedirs(image_folder)

                image_count += 1
                image_filename = f"image_{image_count}.{image_ext}"
                with open(os.path.join(image_folder, image_filename), "wb") as f:
                    f.write(image_bytes)

        if image_count > 0:
            print(
                f"    [Images] Extracted {image_count} images (Filtered out {skipped_count} small items)"
            )
        elif skipped_count > 0:
            print(
                f"    [Images] No main images found (Filtered out {skipped_count} small items)"
            )

        doc.close()
    except Exception as e:
        print(f"    [Images Error] Failed to extract images from {pdf_path}: {e}")


def download_pdf(url, folder_path, filename):
    """
    Downloads a PDF from a URL and saves it to the specified folder.
    """
    if not url or url == "No Link Found" or not url.startswith("http"):
        print(f"  [Skipped] No valid URL for {filename}")
        return False

    download_url = get_direct_download_url(url)

    try:
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        file_path = os.path.join(folder_path, filename)

        response = requests.get(download_url, timeout=45, stream=True)
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "").lower()
        if "text/html" in content_type:
            print(
                f"  [Warning] Received HTML instead of PDF for {filename}. Access may be restricted."
            )

        with open(file_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        print(f"  [Success] Downloaded: {file_path}")

        # Extract images immediately after download
        extract_images_from_pdf(file_path, folder_path)

        return True
    except Exception as e:
        print(f"  [Error] Failed to download {url}: {e}")
        return False


def organize_primary_images():
    """
    Copies the first image from each audit folder to the public app directory
    for easy access by the frontend.
    """
    print("\n" + "=" * 50)
    print(f"Organizing primary images to {PRIMARY_IMAGE_DIR}...")

    if os.path.exists(PRIMARY_IMAGE_DIR):
        shutil.rmtree(PRIMARY_IMAGE_DIR)
    os.makedirs(PRIMARY_IMAGE_DIR)

    # List all directories in the download output (excluding the excel file itself)
    for folder_name in os.listdir(OUTPUT_DIR):
        folder_path = os.path.join(OUTPUT_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue

        images_path = os.path.join(folder_path, "images")
        if not os.path.exists(images_path):
            continue

        # Look for image_1 in the images folder
        files = os.listdir(images_path)
        # Sort to ensure image_1 comes first regardless of filesystem order
        files.sort()

        first_image = None
        for f in files:
            if f.startswith("image_1."):
                first_image = f
                break

        if first_image:
            source = os.path.join(images_path, first_image)
            extension = first_image.split(".")[-1]
            destination = os.path.join(PRIMARY_IMAGE_DIR, f"{folder_name}.{extension}")
            shutil.copy2(source, destination)
            print(f"  [Copy] {folder_name} -> {folder_name}.{extension}")


def fetch_and_organize_walk_audits(sheet_id, gid):
    """
    Downloads the Google Sheet, extracts PDF links, and organizes files into folders.
    """
    if os.path.exists(OUTPUT_DIR):
        print(f"Cleaning up existing directory: {OUTPUT_DIR}...")
        shutil.rmtree(OUTPUT_DIR)

    os.makedirs(OUTPUT_DIR)

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid={gid}"
    excel_path = os.path.join(OUTPUT_DIR, "walk_audit_data.xlsx")

    try:
        print(f"Downloading sheet from Google to {excel_path}...")
        response = requests.get(export_url)
        response.raise_for_status()

        with open(excel_path, "wb") as f:
            f.write(response.content)

        wb = load_workbook(excel_path)
        ws = wb.active

        view_col_idx = None
        header_row_num = 2

        for cell in ws[header_row_num]:
            val = str(cell.value).strip().upper() if cell.value else ""
            if val == "VIEW":
                view_col_idx = cell.column
                break

        if not view_col_idx:
            print(f"Could not find 'VIEW' column in row {header_row_num}.")
            return

        print(
            f"Processing (Limit: {FIRST_N_ROWS if FIRST_N_ROWS else 'All'}) | Min Size: {MIN_IMAGE_WIDTH}x{MIN_IMAGE_HEIGHT}"
        )
        print("=" * 50)

        rows_processed = 0
        for row in ws.iter_rows(min_row=header_row_num + 1):
            if FIRST_N_ROWS is not None and rows_processed >= FIRST_N_ROWS:
                print(f"\nReached limit of {FIRST_N_ROWS} rows. Stopping.")
                break

            if all(cell.value is None for cell in row):
                continue

            city_raw = row[0].value
            year_raw = row[1].value

            target_cell = row[view_col_idx - 1]
            link_url = (
                target_cell.hyperlink.target
                if target_cell.hyperlink
                else "No Link Found"
            )

            city_slug = sanitize_folder_name(city_raw)
            year_slug = sanitize_folder_name(year_raw)
            folder_name = f"{city_slug}_{year_slug}"
            target_folder = os.path.join(OUTPUT_DIR, folder_name)

            file_name = f"walk_audit_{folder_name}.pdf"

            print(f"Row {rows_processed + 1}: {city_raw} ({year_raw})")

            download_pdf(link_url, target_folder, file_name)

            rows_processed += 1

        # After all rows are processed, organize the primary images
        organize_primary_images()

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    SHEET_ID = "1-Vxf7AlXk_WJwwYSVy7F28qjxVXQOAmQ-NN0JImx95Y"
    GID = "379989993"
    fetch_and_organize_walk_audits(SHEET_ID, GID)
